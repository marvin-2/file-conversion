# Extra libraries: defusedxml, openpyxl

import xml.etree.ElementTree as ET 
from xml.dom import minidom
import random
import csv
import json
from openpyxl import DEFUSEDXML
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


MIN_RANGE = 10000000000000
MAX_RANGE = 99999999999999


def gen_csv(toPath):
	print("\ngen_csv(" + toPath + ")")
	headings = ["UPC", "Description"]
	brandNames = [
		"Hershey","Snickers","AlmondJoy", "Reese's", "Mounds", "KitKat", 
    "Twix", "MilkyWay", "Crunch", "Musketeers", "PayDay", "ButterFinger"
		]
    
	sizes = ["Mini", "Standard", "King Size"]			  
	f = open(toPath, "w")
	f.write(",".join(headings) + "\n")
	for b in brandNames:
		for s in sizes:
			f.write(str(random.randrange(MIN_RANGE,MAX_RANGE)) + 
			         "," + b + " " + s  + "\n")
	f.close()


def print_csv(fromPath):
	print("\nprint_csv(" + fromPath + ")")
	f = open(fromPath)
	print(f.read())
	f.close()


def print_xlsx(fromPath):
	"""
	Reads xlsx file
	"""
	print("\nread_xlsx(" + fromPath + ")")
	wb = load_workbook(fromPath)
	sheet = wb.active

	for row in sheet.iter_rows():
		cells = []
		for cell in row:
			cells.append(cell.value)
		print(",".join(cells))


def csv_to_xlsx(fromPath, toPath):
	"""
	Converts CSV file to XLSX
	"""
	print("\ncsv_to_xlsx(" + fromPath + "," + toPath + ")")
	with open(fromPath, newline = '') as f:
		reader = csv.reader(f)
		wb = Workbook()
		sheet = wb.active
		row = 0   
		for ln in reader:
			row += 1
			c1 = sheet.cell(row, 1)
			c1.value = ln[0]
			c2 = sheet.cell(row, 2)
			c2.value = ln[1]

		wb.save(toPath)


def csv_to_xlsx_table(fromPath, toPath):
	"""
	Converts CSV file to table in XLSX file
	"""
	print("\ncsv_to_xlsx_table(" + fromPath + "," + toPath + ")")
	with open(fromPath, newline = '') as f:
		reader = csv.reader(f)
		wb = Workbook()
		sheet = wb.active
		row = 0   
		for ln in reader:
			row += 1
			c1 = sheet.cell(row, 1)
			c1.value = ln[0]
			c2 = sheet.cell(row, 2)
			c2.value = ln[1]

	tab = Table(displayName="Table1", ref="A1:B" + str(row))
	# Add a default style with striped rows and banded columns
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
							showLastColumn=False, showRowStripes=True, showColumnStripes=True)
	tab.tableStyleInfo = style

	sheet.add_table(tab)
	wb.save(toPath)


def test_json(toPath):
	py_obj = {"firstName": "Joe", "lastName": "Anders", "phone": "123-456-7890"}
	# py_obj = {"firstName": "Joe", "lastName": "Anders", "phones": [{"home":"123-456-7890"}, {"work": "987-654-3210"}]}

	# Encoding
	json_str = json.dumps(py_obj, indent = 4)
	print()
	print(type(json_str))
	print(json_str)

	# Decoding
	py_obj2 = json.loads(json_str)
	print()
	print(type(py_obj2))
	print(py_obj2)


def print_json(fromPath):
	with open(fromPath) as jsonF:
		data = jsonF.read()

	print(data)
		


def csv_to_json(fromPath, toPath):
	print("\ncsv_to_json(" + fromPath + "," + toPath + ")")
	data = []

	# Read through csv file 
	with open(fromPath, newline = '') as csvF:
		reader = csv.DictReader(csvF)

		# Append rows to list of dictionaries
		for row in reader:
			data.append(row)

	# Convert list of dictionaries to json, and write to file
	with open(toPath, 'w') as jsonF:
		jsonF.write(json.dumps(data, indent=4)) 


def defused():
	"""
	Installing the defusedxml library helps to safeguard against xml attacks.
	Used by openpyxl, etc.
	"""
	if DEFUSEDXML is True:
		print("Defused")
	else:
		print("Not defused")


def csv_to_xml(toPath):
	# This is the parent (root) tag  
	# onto which other tags would be 
	# created 
	data = ET.Element('chess') 
		
	# Adding a subtag named `Opening` 
	# inside our root tag 
	element1 = ET.SubElement(data, 'Opening') 
		
	# Adding subtags under the `Opening` 
	# subtag  
	s_elem1 = ET.SubElement(element1, 'E4') 
	s_elem2 = ET.SubElement(element1, 'D4') 
		
	# Adding attributes to the tags under 
	# `items` 
	s_elem1.set('type', 'Accepted') 
	s_elem2.set('type', 'Declined') 
		
	# Adding text between the `E4` and `D5`  
	# subtag 
	s_elem1.text = "King's Gambit Accepted"
	s_elem2.text = "Queen's Gambit Declined"
		
	# Converting the xml data to byte object, 
	# for allowing flushing data to file  
	# stream 
	b_xml = ET.tostring(data)
	xmlstr = minidom.parseString(ET.tostring(data)).toprettyxml(indent="   ")
		
	# Opening a file under the name `items2.xml`, 
	# with operation mode `wb` (write + binary) 
	with open(toPath, "wb") as f:
		# f.write(b_xml)
		f.write(xmlstr.encode('utf-8'))
